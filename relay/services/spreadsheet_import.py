"""Import drivers and trucks from CSV or Excel spreadsheets."""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_date

from drivers.models import Driver, DriverStatus, DriverType
from relay.models import AssignmentStatus, RelayAssignment
from relay.services import relay_service
from trucks.models import Truck, TruckStatus

MAX_IMPORT_ROWS = 2000
MAX_ERROR_MESSAGES = 25
INITIAL_ASSIGNMENT_LOOKBACK_WEEKS = 2

DRIVER_REQUIRED = ("first_name", "last_name")
DRIVER_OPTIONAL = ("driver_id", "phone", "email", "status", "driver_type", "hire_date", "notes")
DRIVER_HEADERS = (
    "driver_id",
    "first_name",
    "last_name",
    "phone",
    "email",
    "status",
    "driver_type",
    "hire_date",
    "notes",
)

TRUCK_REQUIRED = ("unit_number",)
TRUCK_OPTIONAL = ("vin", "make", "model", "year", "status", "notes", "driver_id")
TRUCK_HEADERS = (
    "unit_number",
    "driver_id",
    "vin",
    "make",
    "model",
    "year",
    "status",
    "notes",
)

HEADER_ALIASES = {
    "firstname": "first_name",
    "first": "first_name",
    "lastname": "last_name",
    "last": "last_name",
    "telephone": "phone",
    "mobile": "phone",
    "e_mail": "email",
    "mail": "email",
    "driver_status": "status",
    "type": "driver_type",
    "drivertype": "driver_type",
    "driver_type": "driver_type",
    "hiredate": "hire_date",
    "hired": "hire_date",
    "hired_date": "hire_date",
    "start_date": "hire_date",
    "note": "notes",
    "comment": "notes",
    "comments": "notes",
    "unit": "unit_number",
    "unit_no": "unit_number",
    "unitnumber": "unit_number",
    "truck": "unit_number",
    "truck_number": "unit_number",
    "truck_unit": "unit_number",
    "truck_status": "status",
    "full_name": "full_name",
    "name": "full_name",
    "driver_name": "full_name",
    # Canonical column is driver_id; keep legacy aliases.
    "external_id": "driver_id",
    "external_driver_id": "driver_id",
    "source_driver_id": "driver_id",
    "source_id": "driver_id",
    "pt_driver_id": "driver_id",
    "current_driver_id": "driver_id",
    "current_driver_external_id": "driver_id",
    "assigned_driver_id": "driver_id",
}


@dataclass
class SpreadsheetImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    assignments_created: int = 0
    errors: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return self.created > 0 or self.updated > 0 or self.assignments_created > 0


def _normalize_header(raw: Any) -> str:
    text = str(raw or "").strip().lower()
    text = text.replace("-", "_").replace(" ", "_")
    text = re.sub(r"_+", "_", text).strip("_")
    return HEADER_ALIASES.get(text, text)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_date_value(raw: Any) -> date | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    text = _cell_str(raw)
    if not text:
        return None
    parsed = parse_date(text)
    if parsed:
        return parsed
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date: {text}")


def _resolve_choice(raw: str, choices, *, field_label: str) -> str:
    text = (raw or "").strip()
    lowered = text.lower().replace(" ", "_").replace("-", "_")
    valid = {value: value for value, _ in choices}
    if lowered in valid:
        return lowered
    by_label = {label.lower(): value for value, label in choices}
    if text.lower() in by_label:
        return by_label[text.lower()]
    allowed = ", ".join(v for v, _ in choices)
    raise ValueError(f"Invalid {field_label} '{text}'. Use one of: {allowed}")


def _split_full_name(full_name: str) -> tuple[str, str]:
    parts = full_name.strip().split()
    if len(parts) == 1:
        return parts[0], parts[0]
    return parts[0], " ".join(parts[1:])


def read_spreadsheet_rows(uploaded_file) -> list[dict[str, Any]]:
    """Return list of row dicts keyed by normalized header names."""
    name = (getattr(uploaded_file, "name", "") or "").lower()
    raw = uploaded_file.read()
    if not raw:
        raise ValueError("File is empty.")

    if name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return _read_excel_rows(raw)
    if name.endswith(".xls"):
        raise ValueError("Old .xls format is not supported. Save as .xlsx or CSV.")
    return _read_csv_rows(raw)


def _read_csv_rows(raw: bytes) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Could not decode CSV file. Use UTF-8 encoding.")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = list(reader)
    if not rows:
        raise ValueError("CSV has no rows.")
    return _rows_from_matrix(rows)


def _read_excel_rows(raw: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Excel support requires openpyxl. Install project requirements.") from exc

    workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        matrix: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            matrix.append(list(row))
    finally:
        workbook.close()

    if not matrix:
        raise ValueError("Excel sheet is empty.")
    return _rows_from_matrix(matrix)


def _rows_from_matrix(matrix: list[list[Any]]) -> list[dict[str, Any]]:
    header_idx = None
    headers: list[str] = []
    for idx, row in enumerate(matrix):
        candidates = [_normalize_header(cell) for cell in row]
        if any(candidates):
            header_idx = idx
            headers = candidates
            break
    if header_idx is None:
        raise ValueError("Could not find a header row.")

    result: list[dict[str, Any]] = []
    for row in matrix[header_idx + 1 :]:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        mapped: dict[str, Any] = {}
        for col, header in enumerate(headers):
            if not header:
                continue
            value = row[col] if col < len(row) else None
            if header in mapped and mapped[header] not in (None, ""):
                continue
            mapped[header] = value
        result.append(mapped)
        if len(result) > MAX_IMPORT_ROWS:
            raise ValueError(f"Too many rows (max {MAX_IMPORT_ROWS}).")
    if not result:
        raise ValueError("No data rows found under the header.")
    return result


def _add_error(result: SpreadsheetImportResult, message: str) -> None:
    if len(result.errors) < MAX_ERROR_MESSAGES:
        result.errors.append(message)
    elif len(result.errors) == MAX_ERROR_MESSAGES:
        result.errors.append("Additional errors omitted.")


def _find_existing_driver(data: dict[str, Any]) -> Driver | None:
    driver_id = data.get("driver_id")
    if driver_id:
        by_id = Driver.objects.filter(driver_id__iexact=driver_id).order_by("id").first()
        if by_id:
            return by_id
    return (
        Driver.objects.filter(
            first_name__iexact=data["first_name"],
            last_name__iexact=data["last_name"],
        )
        .order_by("id")
        .first()
    )


def import_drivers(rows: list[dict[str, Any]]) -> SpreadsheetImportResult:
    result = SpreadsheetImportResult()
    for index, row in enumerate(rows, start=2):
        try:
            data = _normalize_driver_row(row)
            with transaction.atomic():
                existing = _find_existing_driver(data)
                if existing:
                    for key, value in data.items():
                        if key == "driver_id" and not value and existing.driver_id:
                            continue
                        setattr(existing, key, value)
                    existing.save()
                    result.updated += 1
                else:
                    Driver.objects.create(**data)
                    result.created += 1
        except Exception as exc:  # noqa: BLE001 — collect row errors for user report
            result.skipped += 1
            _add_error(result, f"Row {index}: {exc}")
    return result


def import_trucks(
    rows: list[dict[str, Any]],
    *,
    create_initial_assignments: bool = True,
) -> SpreadsheetImportResult:
    result = SpreadsheetImportResult()
    for index, row in enumerate(rows, start=2):
        try:
            data = _normalize_truck_row(row)
            unit = data["unit_number"]
            driver_ref = data.pop("_link_driver_id", None)
            with transaction.atomic():
                current_driver = None
                if driver_ref:
                    current_driver = _resolve_driver_by_driver_id(driver_ref)
                existing = Truck.objects.filter(unit_number__iexact=unit).first()
                if existing:
                    for key, value in data.items():
                        if key == "unit_number":
                            existing.unit_number = unit
                            continue
                        setattr(existing, key, value)
                    if driver_ref:
                        existing.current_driver = current_driver
                    existing.save()
                    truck = existing
                    result.updated += 1
                else:
                    truck = Truck.objects.create(
                        **data,
                        current_driver=current_driver,
                    )
                    result.created += 1

                if create_initial_assignments and current_driver is not None:
                    if _ensure_initial_assignment(truck, current_driver):
                        result.assignments_created += 1
        except Exception as exc:  # noqa: BLE001
            result.skipped += 1
            _add_error(result, f"Row {index}: {exc}")
    return result


def _truck_has_open_assignment(truck: Truck) -> bool:
    return RelayAssignment.objects.filter(
        truck=truck,
        status__in={AssignmentStatus.ACTIVE, AssignmentStatus.PLANNED},
    ).exists()


def _ensure_initial_assignment(truck: Truck, driver: Driver) -> bool:
    """
    Bootstrap an ACTIVE assignment for initial fleet import.

    Start date is 2 weeks before today so the truck appears mid-cycle
    (default 4-week OTR → home time ~2 weeks ahead). Skips if an open
    assignment already exists — safe for re-import.
    """
    if _truck_has_open_assignment(truck):
        return False

    today = timezone.localdate()
    start_date = today - timedelta(weeks=INITIAL_ASSIGNMENT_LOOKBACK_WEEKS)
    relay_service.create_assignment(
        driver=driver,
        truck=truck,
        start_date=start_date,
        status=AssignmentStatus.ACTIVE,
        notes="Initial assignment from truck import",
    )
    return True


def _normalize_driver_row(row: dict[str, Any]) -> dict[str, Any]:
    first = _cell_str(row.get("first_name"))
    last = _cell_str(row.get("last_name"))
    if (not first or not last) and _cell_str(row.get("full_name")):
        first, last = _split_full_name(_cell_str(row.get("full_name")))
    if not first or not last:
        raise ValueError("first_name and last_name are required (or a full_name column).")

    status_raw = _cell_str(row.get("status"))
    status = (
        _resolve_choice(status_raw, DriverStatus.choices, field_label="status")
        if status_raw
        else DriverStatus.PENDING
    )
    type_raw = _cell_str(row.get("driver_type"))
    driver_type = (
        _resolve_choice(type_raw, DriverType.choices, field_label="driver_type")
        if type_raw
        else DriverType.COMPANY_DRIVER
    )

    email = _cell_str(row.get("email"))
    if email:
        try:
            validate_email(email)
        except ValidationError as exc:
            raise ValueError(f"Invalid email: {email}") from exc

    driver_id = _cell_str(row.get("driver_id"))[:64] or None

    return {
        "driver_id": driver_id,
        "first_name": first[:100],
        "last_name": last[:100],
        "phone": _cell_str(row.get("phone"))[:30],
        "email": email[:254],
        "status": status,
        "driver_type": driver_type,
        "hire_date": _parse_date_value(row.get("hire_date")),
        "notes": _cell_str(row.get("notes")),
    }


def _resolve_driver_by_driver_id(driver_id: str) -> Driver:
    driver = Driver.objects.filter(driver_id__iexact=driver_id).order_by("id").first()
    if not driver:
        raise ValueError(f"No driver found with Driver ID '{driver_id}'.")
    return driver


def _normalize_truck_row(row: dict[str, Any]) -> dict[str, Any]:
    unit = _cell_str(row.get("unit_number"))
    if not unit:
        raise ValueError("unit_number is required.")

    status_raw = _cell_str(row.get("status"))
    status = (
        _resolve_choice(status_raw, TruckStatus.choices, field_label="status")
        if status_raw
        else TruckStatus.AVAILABLE
    )

    year_raw = row.get("year")
    year = None
    if year_raw is not None and _cell_str(year_raw) != "":
        try:
            year = int(float(_cell_str(year_raw)))
        except ValueError as exc:
            raise ValueError(f"Invalid year: {year_raw}") from exc
        if year < 1990 or year > 2100:
            raise ValueError(f"Year out of range: {year}")

    driver_ref = _cell_str(row.get("driver_id"))[:64] or None

    return {
        "unit_number": unit[:50],
        "vin": _cell_str(row.get("vin"))[:17],
        "make": _cell_str(row.get("make"))[:100],
        "model": _cell_str(row.get("model"))[:100],
        "year": year,
        "status": status,
        "notes": _cell_str(row.get("notes")),
        "_link_driver_id": driver_ref,
    }


def build_template_csv(kind: str) -> str:
    if kind == "drivers":
        headers = list(DRIVER_HEADERS)
        sample = [
            "DRV-1001",
            "John",
            "Smith",
            "555-0100",
            "john@example.com",
            "active",
            "company_driver",
            "2024-01-15",
            "",
        ]
    elif kind == "trucks":
        headers = list(TRUCK_HEADERS)
        sample = ["T-101", "DRV-1001", "", "Freightliner", "Cascadia", "2022", "available", ""]
    else:
        raise ValueError("Unknown template kind.")

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerow(sample)
    return buffer.getvalue()
